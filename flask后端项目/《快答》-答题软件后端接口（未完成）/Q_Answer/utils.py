"""
作者：zwq
日期:2022年08月18日
"""
import openpyxl

""" 工具函数 """


class Excel:
    """excel 相关操作"""
    def __init__(self, path):
        # 打开表
        self.excel = openpyxl.load_workbook(path)
        # 选择页数为第1页
        self.sheet = self.excel["Sheet1"]

    def read(self):
        """读取上传文件"""
        # 数据总行数
        nrows = self.sheet.max_row
        info = []
        for i in range(2,nrows+1):
            dicts = {
                "subject": self.sheet.cell(row=i, column=1).value,
                "answer": self.sheet.cell(row=i, column=2).value,
                "answerTrue": self.sheet.cell(row=i, column=3).value,
                "analysis": self.sheet.cell(row=i, column=4).value,
                "classification": self.sheet.cell(row=i, column=5).value
            }
            info.append(dicts)
        return info


if __name__ == '__main__':
    data = Excel(r"D:\PycharmProjects\webProject\project\Q_Answer\upload\2022-08-19\1660655347\template.xlsx").read()
    print(data)
