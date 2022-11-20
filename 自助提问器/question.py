"""
作者：zwq
日期:2022年08月22日
"""

import pyttsx3
import openpyxl
import random

class AskQuestion():
    """ 语音播报类 """

    def __init__(self):
        self.engine = pyttsx3.init()  # 初始化语音引擎

    def setRate(self, rate):
        """设置语速"""
        self.engine.setProperty('rate', rate)  # 设置语速

    def setVolume(self, volume):
        """ 设置音量 """
        self.engine.setProperty('volume', volume)  # 设置音量 0-1

    def speak(self, question):
        self.engine.say(question)
        self.engine.runAndWait()
        self.engine.stop()


class Excel:
    """excel 相关操作 读取问题"""

    def __init__(self,path):
        # 打开表
        self.excel = openpyxl.load_workbook(path)
        # 选择页数为第1页
        self.sheet = self.excel["Sheet1"]

    def read(self):
        """读取上传文件"""
        # 数据总行数
        nrows = self.sheet.max_row
        info = []
        for i in range(2, nrows + 1):
            item = self.sheet.cell(row=i, column=1).value
            info.append(item)
        return info


class question:
    def __init__(self,path):
        excel = Excel(path)
        self.questionList = excel.read()

    def getquestion(self):
        return random.choice(self.questionList)






if __name__ == '__main__':
    # 实例化播报类
    bobao = AskQuestion()
    # 实例化 问题类
    ceshi = question()


    question = ceshi.getquestion()
    bobao.speak(question)